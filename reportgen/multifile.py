"""'하루 1파일'로 쌓인 원본을 모아 월간표(하루 1행) 하나로 만든다.

일부 설비 로깅 프로그램(공조기, 냉동기, 유량계 등)은 매일 새 엑셀 파일을
하나씩 만든다 — 그 안에는 그날의 시간별 데이터와 맨 아래 그날 요약값(일
사용량 등)이 들어 있다. 월간 보고서를 만들려면 그런 파일을 그 달치만큼
모아서, 파일마다 '그날의 요약 행'을 하나씩 뽑아 날짜순으로 쌓아야 한다.

설비마다 양식이 다른 것도 (공조기 일지와 냉동기 일지는 시트 이름도 표
위치도 다르다) **한 번에 월간표 하나로** 합칠 수 있다: 설비별로 읽는 방법을
:class:`SourceSpec` 하나씩 정의해 :func:`build_combined_monthly_table` 에
넘기면, 각 설비에서 뽑은 하루치 값을 **날짜를 기준으로 한 줄에 나란히**
붙인다. 컬럼 이름은 설비 이름을 앞에 붙여(``'공조기 · 가동시간'``) 서로
겹치지 않게 한다.

이 모듈은 그 취합만 담당한다. 결과는 보통의 :class:`~reportgen.data_reader.Table`
과 똑같은 모양(컬럼 + 행)으로 나오므로, 그 뒤로는 기존 매핑/생성 파이프라인을
그대로 쓸 수 있다 — 실제로 GUI 는 이 결과를 엑셀 파일로 저장해서, '1단계
원본 엑셀'에 그대로 넣어 이어서 쓰게 한다.
"""

from __future__ import annotations

import datetime as _dt
import os
import re
from dataclasses import dataclass, field
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

from .data_reader import ReadOptions, Table, read_table
from .dateutils import parse_date
from .errors import MultiFileError, ReportGenError

__all__ = [
    "DailyRowSpec",
    "DailyExtraction",
    "SourceSpec",
    "list_daily_files",
    "extract_date",
    "preview_daily_file",
    "build_monthly_table",
    "build_combined_monthly_table",
    "save_table_as_excel",
]

_EXTENSIONS = (".xlsx", ".xlsm")
#: 파일명 안의 '20260801' 형태 날짜
_RE_FILENAME_DATE = re.compile(r"(20\d{2})[-_.]?(\d{2})[-_.]?(\d{2})")
#: 셀 안의 날짜를 찾을 때 훑어보는 범위 (제목 줄이 보통 위쪽에 있으므로 작게 잡는다)
_SCAN_ROWS = 15
_SCAN_COLS = 12
#: 연도로 보일 만한 4자리 숫자가 있어야 '날짜 같은 문자열'로 본다.
_RE_LOOKS_LIKE_YEAR = re.compile(r"(19|20)\d{2}")


@dataclass
class DailyRowSpec:
    """파일 하나에서 '그날의 요약'으로 뽑아낼 행을 지정한다.

    ``row_indexes`` 는 :func:`~reportgen.data_reader.read_table` 로 읽은
    표의 0부터 시작하는 행 번호다(헤더 제외, 데이터 행 기준). 두 개 이상
    지정하면 한 파일에서 여러 줄을 뽑아 한 줄로 합친다 — 예를 들어 '일사용량
    (TON)' 행과 '일사용량 (N/M3)' 행이 따로 떨어져 있을 때 둘 다 필요하면
    이렇게 쓴다. 이때 ``row_labels`` 를 같이 주면 합쳐질 때 컬럼 이름 뒤에
    ``' (라벨)'`` 을 붙여 구분한다(비우면 나중 행이 앞 행의 같은 이름 값을
    덮어쓸 수 있으니 권장).
    """

    row_indexes: list[int] = field(default_factory=list)
    row_labels: list[str] = field(default_factory=list)
    #: 비우면 표의 모든 컬럼을 쓴다.
    columns: list[str] = field(default_factory=list)


@dataclass
class DailyExtraction:
    files: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def list_daily_files(folder: str) -> list[str]:
    """폴더에서 날짜별 원본 파일 후보를 찾는다 (엑셀이 만드는 임시 잠금 파일 제외)."""
    if not os.path.isdir(folder):
        raise MultiFileError(
            f"폴더를 찾을 수 없습니다: {folder}",
            "폴더 경로를 다시 확인해 주세요.",
        )
    names = []
    for name in os.listdir(folder):
        if name.startswith("~$"):
            continue
        if os.path.splitext(name)[1].lower() not in _EXTENSIONS:
            continue
        names.append(name)
    return sorted(os.path.join(folder, name) for name in names)


def extract_date(path: str) -> Optional[_dt.date]:
    """파일 하나의 날짜를 추정한다.

    1) 파일명 안의 '20260801' / '2026-08-01' 형태를 먼저 본다(가장 흔하고
       빠르다).
    2) 못 찾으면 파일을 열어 위쪽 몇 줄·몇 칸 안에서 날짜로 읽히는 셀을
       찾는다(예: 제목 줄의 '2026년 8월 1일 토요일').
    """
    stem = os.path.splitext(os.path.basename(path))[0]
    match = _RE_FILENAME_DATE.search(stem)
    if match:
        year, month, day = (int(g) for g in match.groups())
        try:
            return _dt.date(year, month, day)
        except ValueError:
            pass

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001 - 못 열면 그냥 못 찾은 것으로 취급
        return None
    try:
        sheet = workbook[workbook.sheetnames[0]]
        max_row = min(sheet.max_row or 0, _SCAN_ROWS)
        max_col = min(sheet.max_column or 0, _SCAN_COLS)
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                found = _looks_like_date_cell(cell.value)
                if found is not None:
                    return found
        return None
    finally:
        workbook.close()


def _looks_like_date_cell(value: Any) -> Optional[_dt.date]:
    """제목 줄 등에서 '날짜로 보이는 셀'만 조심스럽게 골라낸다.

    ``parse_date`` 를 아무 셀에나 그대로 돌리면 위험하다 — 예를 들어 작은
    숫자(1, 10 같은 시간·수량 값)도 엑셀 일련번호로 해석돼 버려서, 데이터
    칸을 날짜로 잘못 인식할 수 있다. 그래서 여기서는 (1) 이미 날짜/시간
    형식으로 읽힌 값이거나 (2) 4자리 연도(19xx/20xx)가 들어 있는 문자열만
    후보로 본다.
    """
    if isinstance(value, (_dt.datetime, _dt.date)):
        return parse_date(value)
    if isinstance(value, str):
        text = value.strip()
        if not text or not _RE_LOOKS_LIKE_YEAR.search(text):
            return None
        return parse_date(text)
    return None


def preview_daily_file(path: str, options: Optional[ReadOptions] = None) -> Table:
    """샘플 파일 하나를 읽어 표로 보여준다 — GUI 가 '몇 번째 행을 가져올지' 고를 때 쓴다."""
    return read_table(path, options)


@dataclass
class SourceSpec:
    """설비 하나(폴더 하나)를 읽는 방법.

    설비마다 양식이 다르므로 — 공조기 일지와 냉동기 일지는 시트 이름도 표
    위치도 다르다 — 폴더별로 이 묶음을 하나씩 만들어
    :func:`build_combined_monthly_table` 에 넘긴다.

    ``name`` 은 합쳐진 표에서 컬럼 이름 앞에 붙는다(``'공조기 · 가동시간'``).
    설비끼리 컬럼 이름이 같아도(둘 다 '가동시간') 섞이지 않게 하기 위한 것이라
    비워 두면 안 된다.
    """

    name: str
    folder: str
    spec: DailyRowSpec
    options: Optional[ReadOptions] = None


def _extract_by_date(
    folder: str, spec: DailyRowSpec, options: Optional[ReadOptions]
) -> tuple[dict[_dt.date, dict[str, Any]], list[str], list[str]]:
    """폴더 하나에서 ``{날짜: {컬럼: 값}}`` 과 컬럼 순서, 경고를 뽑아낸다."""
    if not spec.row_indexes:
        raise MultiFileError(
            "가져올 행 번호를 하나 이상 지정해 주세요.",
            "미리보기에서 그날 요약값이 있는 행 번호를 확인해 입력해 주세요.",
        )

    files = list_daily_files(folder)
    if not files:
        raise MultiFileError(
            f"'{folder}' 폴더에서 엑셀 파일을 찾지 못했습니다.",
            "폴더 안에 .xlsx / .xlsm 파일이 있는지 확인해 주세요.",
        )

    warnings: list[str] = []
    by_date: dict[_dt.date, dict[str, Any]] = {}
    columns: list[str] = []

    for path in files:
        name = os.path.basename(path)
        day = extract_date(path)
        if day is None:
            warnings.append(f"{name}: 날짜를 찾지 못해 건너뜀 (파일명이나 셀에서 날짜를 못 읽음)")
            continue

        try:
            table = read_table(path, options)
        except ReportGenError as exc:
            warnings.append(f"{name}: 읽지 못해 건너뜀 ({exc.message})")
            continue

        row_dict: dict[str, Any] = {}
        multi = len(spec.row_indexes) > 1 and len(spec.row_labels) == len(spec.row_indexes)
        skipped_rows = []
        for position, idx in enumerate(spec.row_indexes):
            if idx < 0 or idx >= len(table.rows):
                skipped_rows.append(idx)
                continue
            row = table.rows[idx]
            wanted = spec.columns or table.columns
            tag = f" ({spec.row_labels[position]})" if multi else ""
            for col_name in wanted:
                if col_name not in table.columns:
                    continue
                value = row[table.index_of(col_name)]
                key = f"{col_name}{tag}"
                row_dict[key] = value
                if key not in columns:
                    columns.append(key)
        if skipped_rows:
            warnings.append(
                f"{name}: 지정한 행 번호 {skipped_rows} 이(가) 이 파일에는 없음 "
                f"(이 파일은 데이터 행이 {len(table.rows)}개뿐)"
            )
        if not row_dict:
            warnings.append(f"{name}: 지정한 행에서 값을 하나도 못 뽑아 건너뜀")
            continue

        if day in by_date:
            warnings.append(f"{name}: 같은 날짜({day.isoformat()})의 다른 파일과 겹쳐 나중 파일로 덮어씀")
        by_date[day] = row_dict

    return by_date, columns, warnings


def build_monthly_table(
    folder: str,
    spec: DailyRowSpec,
    options: Optional[ReadOptions] = None,
    date_column_name: str = "날짜",
) -> tuple[Table, list[str]]:
    """폴더 안의 날짜별 파일을 전부 읽어, 파일마다 지정한 행을 뽑아 한 달치 표로 합친다.

    돌려주는 두 번째 값(경고 목록)에는 날짜를 못 찾았거나, 지정한 행 번호가
    그 파일에 없거나, 읽는 중 오류가 난 파일이 무엇이었는지 남긴다 — 원본을
    직접 보지 않고도 어떤 날짜가 왜 빠졌는지 알 수 있게 하기 위해서다.
    """
    by_date, columns, warnings = _extract_by_date(folder, spec, options)

    if not by_date:
        raise MultiFileError(
            "폴더의 파일에서 월간표에 넣을 값을 하나도 뽑지 못했습니다.",
            "행 번호가 맞는지, 날짜를 인식하는지 미리보기로 다시 확인해 주세요.",
        )

    ordered_days = sorted(by_date)
    all_columns = [date_column_name] + columns
    rows: list[list[Any]] = []
    for day in ordered_days:
        values = by_date[day]
        rows.append([day] + [values.get(col) for col in columns])

    table = Table(columns=all_columns, rows=rows, sheet_name="월간취합", source_path=folder)
    return table, warnings


def build_combined_monthly_table(
    sources: list[SourceSpec],
    date_column_name: str = "날짜",
) -> tuple[Table, list[str]]:
    """여러 설비(폴더)를 **한 번에 월간표 하나로** 합친다.

    설비마다 양식이 달라도 된다 — 폴더별로 시트/범위/행 번호를 따로 정하기
    때문이다. 각 설비에서 뽑은 하루치 값을 **날짜를 기준으로 한 줄에 나란히**
    붙이고, 컬럼 이름 앞에 설비 이름을 붙여(``'공조기 · 가동시간'``) 설비끼리
    이름이 같아도 섞이지 않게 한다.

    어떤 설비에 그 날짜 파일이 없으면 그 칸만 비워 둔다(그 날짜 줄 전체가
    빠지지는 않는다) — 설비마다 기록이 시작된 날이 다를 수 있기 때문이다.
    빠진 칸은 경고로 남겨 어떤 설비의 며칠 치가 없는지 알 수 있게 한다.
    """
    if not sources:
        raise MultiFileError(
            "합칠 원본을 하나 이상 추가해 주세요.",
            "설비(폴더)마다 한 줄씩 추가한 뒤 다시 시도해 주세요.",
        )

    seen_names: set[str] = set()
    for source in sources:
        if not source.name.strip():
            raise MultiFileError(
                "원본 이름이 비어 있습니다.",
                "합쳐진 표에서 컬럼을 구분하려면 설비 이름(예: 공조기, 냉동기)이 필요합니다.",
            )
        if source.name in seen_names:
            raise MultiFileError(
                f"원본 이름 '{source.name}' 이(가) 중복됩니다.",
                "설비 이름은 서로 달라야 컬럼이 섞이지 않습니다.",
            )
        seen_names.add(source.name)

    warnings: list[str] = []
    merged: dict[_dt.date, dict[str, Any]] = {}
    all_columns: list[str] = []
    days_by_source: dict[str, set[_dt.date]] = {}

    for source in sources:
        try:
            by_date, columns, source_warnings = _extract_by_date(
                source.folder, source.spec, source.options
            )
        except MultiFileError as exc:
            # 설비 하나가 실패해도 나머지는 살린다 — 한 폴더 때문에 전체가
            # 날아가면 '한 번에 하나로 합치기'가 무의미해진다.
            warnings.append(f"[{source.name}] 건너뜀: {exc.message}")
            continue

        warnings.extend(f"[{source.name}] {w}" for w in source_warnings)
        if not by_date:
            warnings.append(f"[{source.name}] 이 폴더에서는 값을 하나도 뽑지 못해 빠졌습니다.")
            continue

        days_by_source[source.name] = set(by_date)
        prefix = f"{source.name} · "
        for column in columns:
            key = f"{prefix}{column}"
            if key not in all_columns:
                all_columns.append(key)
        for day, values in by_date.items():
            slot = merged.setdefault(day, {})
            for column, value in values.items():
                slot[f"{prefix}{column}"] = value

    if not merged:
        raise MultiFileError(
            "어느 폴더에서도 월간표에 넣을 값을 뽑지 못했습니다.",
            "폴더 경로·행 번호·시트 범위가 맞는지 미리보기로 다시 확인해 주세요.",
        )

    ordered_days = sorted(merged)
    # 설비마다 비는 날짜가 있으면 알려 준다 (그 줄 전체를 버리지는 않는다).
    for name, days in days_by_source.items():
        missing = [d for d in ordered_days if d not in days]
        if missing:
            preview = ", ".join(d.isoformat() for d in missing[:5])
            more = f" 외 {len(missing) - 5}일" if len(missing) > 5 else ""
            warnings.append(f"[{name}] 다른 설비엔 있는데 이 설비엔 없는 날짜 {len(missing)}일: {preview}{more}")

    rows: list[list[Any]] = []
    for day in ordered_days:
        values = merged[day]
        rows.append([day] + [values.get(col) for col in all_columns])

    table = Table(
        columns=[date_column_name] + all_columns,
        rows=rows,
        sheet_name="월간취합",
        source_path=sources[0].folder,
    )
    return table, warnings


def save_table_as_excel(table: Table, path: str) -> str:
    """취합한 표를 실제 .xlsx 파일로 저장한다 (1단계 '원본 엑셀'에 그대로 넣어 쓸 수 있게)."""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (table.sheet_name or "월간취합")[:31] or "월간취합"
    sheet.append(list(table.columns))
    for row in table.rows:
        sheet.append([_coerce(v) for v in row])
    workbook.save(path)
    return os.path.abspath(path)


def _coerce(value: Any) -> Any:
    if isinstance(value, (int, float, bool, str, _dt.date, _dt.datetime, type(None))):
        return value
    return str(value)
