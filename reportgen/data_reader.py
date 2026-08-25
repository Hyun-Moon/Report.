"""원본 엑셀(.xlsx) 읽기.

openpyxl 만 사용한다(pandas 불필요). 결과는 :class:`Table` 하나로 통일되며
이후 집계/매핑 단계는 이 자료구조만 알면 된다.

지원하는 형태
-------------
* 헤더 1행 (가장 흔함)
* 헤더 여러 행 (병합된 상위 헤더 + 하위 헤더) -> ``'상위 / 하위'`` 로 합침
* 표가 시트 왼쪽 위에서 시작하지 않는 경우 -> 셀 범위 지정 또는 자동 탐색
* 날짜가 '열' 방향으로 늘어선 표 -> ``transpose=True``
"""

from __future__ import annotations

import datetime as _dt
import os
import re
from dataclasses import dataclass, field
from typing import Any, Iterable, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

from .errors import (
    CellRangeError,
    FileFormatError,
    FormulaCacheError,
    HeaderError,
    SheetNotFoundError,
)

__all__ = [
    "Table",
    "ReadOptions",
    "TableBlock",
    "read_table",
    "list_sheets",
    "parse_range",
    "list_table_blocks",
    "FormulaCacheError",
]

_RE_RANGE = re.compile(r"^\s*([A-Za-z]{1,3})(\d+)\s*:\s*([A-Za-z]{1,3})(\d+)\s*$")
#: 표의 '시작 위치와 컬럼 범위'를 찾을 때만 훑어보는 창. 데이터 끝 행은 이 값과
#: 무관하게 시트 전체를 기준으로 잡는다(긴 표가 잘리면 안 되므로).
_SCAN_ROWS = 300
_SCAN_COLS = 100


# --------------------------------------------------------------------------- #
# 자료구조
# --------------------------------------------------------------------------- #
@dataclass
class Table:
    """헤더 + 데이터 행으로 이루어진 2차원 표."""

    columns: list[str]
    rows: list[list[Any]] = field(default_factory=list)
    sheet_name: str = ""
    source_path: str = ""
    #: 읽는 과정에서 사용자에게 알려야 할 안내 (예: 계산 안 된 수식을 빈 값으로 넘김)
    warnings: list[str] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.columns = _dedupe(self.columns)

    # -- 조회 -------------------------------------------------------------- #
    def index_of(self, column: str) -> int:
        try:
            return self.columns.index(column)
        except ValueError as exc:
            raise KeyError(column) from exc

    def column_values(self, column: str) -> list[Any]:
        idx = self.index_of(column)
        return [row[idx] if idx < len(row) else None for row in self.rows]

    def cell(self, row_index: int, column: str) -> Any:
        idx = self.index_of(column)
        row = self.rows[row_index]
        return row[idx] if idx < len(row) else None

    def row_dicts(self) -> list[dict[str, Any]]:
        """docxtpl 반복문(``{% for r in rows %}``)에 넘기기 좋은 형태."""
        out: list[dict[str, Any]] = []
        for row in self.rows:
            item: dict[str, Any] = {}
            for i, name in enumerate(self.columns):
                item[name] = row[i] if i < len(row) else None
            out.append(item)
        return out

    @property
    def n_rows(self) -> int:
        return len(self.rows)

    def preview(self, limit: int = 20) -> list[list[Any]]:
        return [list(self.columns)] + [list(r) for r in self.rows[:limit]]

    def transposed(self) -> "Table":
        """행/열을 뒤집는다. 첫 컬럼의 값들이 새 헤더가 된다."""
        if not self.columns:
            return Table([], [], self.sheet_name, self.source_path)
        new_columns = [str(self.columns[0])] + [
            _stringify(row[0]) for row in self.rows
        ]
        new_rows: list[list[Any]] = []
        for col_idx in range(1, len(self.columns)):
            new_row: list[Any] = [self.columns[col_idx]]
            for row in self.rows:
                new_row.append(row[col_idx] if col_idx < len(row) else None)
            new_rows.append(new_row)
        return Table(new_columns, new_rows, self.sheet_name, self.source_path)


@dataclass
class ReadOptions:
    """GUI 1단계에서 사용자가 고르는 값들."""

    sheet_name: Optional[str] = None
    cell_range: Optional[str] = None
    header_rows: int = 1
    auto_detect: bool = True
    transpose: bool = False
    skip_blank_rows: bool = True
    #: True 면 계산되지 않은 수식 셀을 오류 대신 빈 값으로 처리하고 넘어간다.
    #: (원본을 엑셀로 열어 재계산할 수 없는 상황을 위한 탈출구. 기본은 False —
    #: 그 값들이 조용히 빈칸/0 으로 집계될 수 있으므로 사용자가 직접 켜야 한다.)
    allow_uncalculated_formulas: bool = False


# --------------------------------------------------------------------------- #
# 공개 함수
# --------------------------------------------------------------------------- #
def list_sheets(path: str) -> list[str]:
    """워크북의 시트 이름 목록."""
    workbook = _open_workbook(path, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def parse_range(text: str) -> tuple[int, int, int, int]:
    """``'B3:F40'`` -> ``(min_row, min_col, max_row, max_col)`` (1-based)."""
    match = _RE_RANGE.match(text or "")
    if not match:
        raise CellRangeError(
            f"셀 범위 '{text}' 를 이해할 수 없습니다.",
            "'B3:F40' 처럼 [시작셀]:[끝셀] 형식으로 입력해 주세요.",
        )
    col1 = column_index_from_string(match.group(1).upper())
    row1 = int(match.group(2))
    col2 = column_index_from_string(match.group(3).upper())
    row2 = int(match.group(4))
    if row1 > row2 or col1 > col2:
        row1, row2 = min(row1, row2), max(row1, row2)
        col1, col2 = min(col1, col2), max(col1, col2)
    return row1, col1, row2, col2


@dataclass
class TableBlock:
    """시트 안에서 찾아낸 '표처럼 보이는 덩어리' 하나."""

    range: str  # "B3:N5"
    row1: int
    col1: int
    row2: int
    col2: int
    n_rows: int
    n_cols: int
    preview: str  # 첫 줄 내용 요약 (표를 구분하는 용도)

    def label(self) -> str:
        return f"{self.range}  ({self.n_rows}행 x {self.n_cols}열)  {self.preview}"


def list_table_blocks(path: str, sheet_name: Optional[str] = None) -> list[TableBlock]:
    """한 시트 안에 표가 여러 개 뒤섞여 있어도, 표처럼 보이는 덩어리를 전부 찾는다.

    실무 엑셀은 시트 하나에 표 하나만 깔끔하게 있는 경우가 오히려 드물다 —
    제목 줄, 서로 다른 표 여러 개가 위아래·좌우로 붙어 있는 경우가 흔하다.
    이 함수는 채워진 셀들을 상하좌우로 이어진 덩어리(연결 요소)로 묶어서
    표 후보 목록을 만든다. GUI 1단계의 [표 후보 찾기] 가 이 함수를 쓴다.

    ``자동 감지`` 하나가 시트 전체를 한 표로 오인하는 문제(여러 표가 섞인
    실제 업무 파일에서 흔히 일어남)를, 후보를 보여주고 사람이 고르게 해서
    피해간다.
    """
    workbook = _open_workbook(path, data_only=True)
    try:
        sheet = _pick_sheet(workbook, sheet_name, path)
        merged = _merged_lookup(sheet)
        max_row = min(sheet.max_row or 0, _SCAN_ROWS)
        max_col = min(sheet.max_column or 0, _SCAN_COLS)
        if max_row < 1 or max_col < 1:
            return []

        filled: set[tuple[int, int]] = set()
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                if _is_filled(_value_at(sheet, merged, r, c)):
                    filled.add((r, c))
        if not filled:
            return []

        blocks: list[TableBlock] = []
        for cells in _connected_components(filled):
            rows = [r for r, _ in cells]
            cols = [c for _, c in cells]
            row1, row2 = min(rows), max(rows)
            col1, col2 = min(cols), max(cols)
            if row1 == row2 and col1 == col2:
                continue  # 셀 하나짜리 '섬'(스치는 글자)은 표로 안 본다
            if len(cells) < 3:
                continue

            preview_bits: list[str] = []
            for c in range(col1, min(col2, col1 + 4) + 1):
                value = _value_at(sheet, merged, row1, c)
                text = _stringify(value)
                if text:
                    preview_bits.append(text)
            blocks.append(
                TableBlock(
                    range=f"{get_column_letter(col1)}{row1}:{get_column_letter(col2)}{row2}",
                    row1=row1,
                    col1=col1,
                    row2=row2,
                    col2=col2,
                    n_rows=row2 - row1 + 1,
                    n_cols=col2 - col1 + 1,
                    preview=", ".join(preview_bits) if preview_bits else "(제목 없음)",
                )
            )

        blocks.sort(key=lambda b: (b.row1, b.col1))
        return blocks
    finally:
        workbook.close()


def _connected_components(cells: set[tuple[int, int]]) -> list[list[tuple[int, int]]]:
    """채워진 셀들을 상하좌우로 붙어 있는 덩어리 단위로 묶는다 (너비우선 탐색).

    대각선으로만 붙어 있거나, 빈 행/열 하나로 떨어져 있는 셀은 별도 덩어리로
    본다 — 그래야 옆으로 나란히 놓인 표들과, 위아래로 쌓인 표들이 각각
    구분된다.
    """
    remaining = set(cells)
    components: list[list[tuple[int, int]]] = []
    while remaining:
        start = next(iter(remaining))
        remaining.discard(start)
        queue = [start]
        group = [start]
        while queue:
            r, c = queue.pop()
            for neighbor in ((r - 1, c), (r + 1, c), (r, c - 1), (r, c + 1)):
                if neighbor in remaining:
                    remaining.discard(neighbor)
                    queue.append(neighbor)
                    group.append(neighbor)
        components.append(group)
    return components


def read_table(path: str, options: Optional[ReadOptions] = None) -> Table:
    """원본 엑셀에서 표 하나를 읽어 :class:`Table` 로 돌려준다."""
    options = options or ReadOptions()
    workbook = _open_workbook(path, data_only=True)
    try:
        sheet = _pick_sheet(workbook, options.sheet_name, path)
        merged = _merged_lookup(sheet)

        if options.cell_range:
            row1, col1, row2, col2 = parse_range(options.cell_range)
        else:
            bounds = _detect_bounds(sheet, merged)
            if bounds is None:
                raise HeaderError(
                    f"'{sheet.title}' 시트에서 읽을 수 있는 표를 찾지 못했습니다.",
                    "시트를 다시 고르거나, 셀 범위를 직접 입력해 주세요.",
                )
            row1, col1, row2, col2 = bounds

        grid = _read_grid(sheet, merged, row1, col1, row2, col2)
        if not grid:
            raise HeaderError(
                f"'{sheet.title}' 시트의 지정 범위가 비어 있습니다.",
                "셀 범위를 확인해 주세요.",
            )

        header_rows = max(1, int(options.header_rows or 1))
        if options.auto_detect and not options.cell_range:
            header_rows = _detect_header_rows(grid, header_rows)
        header_rows = min(header_rows, len(grid))

        table_warnings: list[str] = []
        strict = not options.allow_uncalculated_formulas

        # 헤더 행 자체가 계산 안 된 수식이면(예: 하루씩 더해가는 날짜 헤더),
        # 컬럼 이름을 만들기도 전에 걸러야 한다. 그러지 않으면 '열C', '열D'
        # 같은 엉뚱한 이름이 오류 없이 조용히 생겨 버린다.
        if any(v is None for row in grid[:header_rows] for v in row):
            offenders = _find_uncached_formula_cells(
                path, sheet.title, row1, col1, row1 + header_rows - 1, col2
            )
            if offenders:
                table_warnings.extend(
                    _formula_cache_message(offenders, strict, context="헤더 행")
                )

        columns = _build_columns(grid[:header_rows], col1)
        body = grid[header_rows:]

        if any(v is None for row in body for v in row):
            offenders = _find_uncached_formula_cells(
                path, sheet.title, row1 + header_rows, col1, row2, col2
            )
            if offenders:
                table_warnings.extend(
                    _formula_cache_message(offenders, strict, context="데이터", columns=columns, col1=col1)
                )

        if options.skip_blank_rows:
            body = [row for row in body if any(_is_filled(v) for v in row)]

        table = Table(columns, body, sheet.title, os.path.abspath(path), warnings=table_warnings)
        if options.transpose:
            table = table.transposed()
        return table
    finally:
        workbook.close()


# --------------------------------------------------------------------------- #
# 내부 구현
# --------------------------------------------------------------------------- #
def _open_workbook(path: str, data_only: bool):
    if not path:
        raise FileFormatError("엑셀 파일을 먼저 선택해 주세요.")
    if not os.path.isfile(path):
        raise FileFormatError(f"파일을 찾을 수 없습니다: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        raise FileFormatError(
            f"'{ext or '확장자 없음'}' 형식은 읽을 수 없습니다.",
            "구형 .xls 파일은 엑셀에서 .xlsx 로 다시 저장한 뒤 사용해 주세요.",
        )
    try:
        return load_workbook(path, data_only=data_only, read_only=False)
    except Exception as exc:  # openpyxl 은 다양한 예외를 던진다
        raise FileFormatError(
            f"엑셀 파일을 열지 못했습니다: {os.path.basename(path)}",
            f"파일이 손상되었거나 다른 프로그램이 열고 있을 수 있습니다. ({exc})",
        ) from exc


def _find_uncached_formula_cells(
    path: str, sheet_name: str, row1: int, col1: int, row2: int, col2: int
) -> list[tuple[str, int, int]]:
    """지정한 범위 안에서 '계산되지 않은 수식' 셀 좌표를 찾는다.

    엑셀 파일은 수식과 별개로 마지막 계산 결과를 셀에 캐시해 둔다. 이
    프로그램은 그 캐시만 읽으므로(``data_only=True``), 프로그램이 만들었거나
    LibreOffice 등에서 재계산 없이 저장된 파일은 수식 칸이 조용히 빈 값으로
    읽힌다. 집계가 슬쩍 틀어지는 것보다는 여기서 바로 알려주는 편이 안전하다.

    ``(좌표문자열, 행, 열)`` 목록을 돌려준다.
    """
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        sheet = workbook[sheet_name]
        offenders: list[tuple[str, int, int]] = []
        for r in range(row1, row2 + 1):
            for c in range(col1, col2 + 1):
                cell = sheet.cell(row=r, column=c)
                if cell.data_type == "f":
                    offenders.append((cell.coordinate, r, c))
        return offenders
    finally:
        workbook.close()


def _formula_cache_message(
    offenders: list[tuple[str, int, int]],
    strict: bool,
    context: str,
    columns: Optional[list[str]] = None,
    col1: Optional[int] = None,
) -> list[str]:
    """계산 안 된 수식 셀 목록을 사람이 읽을 메시지로 만든다.

    ``strict=True`` 면 :class:`FormulaCacheError` 를 그대로 던진다.
    ``strict=False`` 면 (원본을 다시 계산할 방법이 없는 경우를 위한 탈출구)
    오류 대신 경고 문구 목록을 돌려주고, 그 칸들은 빈 값으로 두고 진행한다 —
    조용히 넘어가지는 않는다.
    """
    if columns is not None and col1 is not None:
        labels = [f"{columns[c - col1]} ({coord})" for coord, _r, c in offenders]
    else:
        labels = [coord for coord, _r, _c in offenders]

    preview = ", ".join(labels[:5])
    more = f" 외 {len(labels) - 5}건" if len(labels) > 5 else ""
    if strict:
        raise FormulaCacheError(
            f"원본 엑셀의 {context}에 계산되지 않은 수식 셀이 있어 값을 읽을 수 없습니다.",
            f"엑셀(또는 한셀/LibreOffice Calc 등)에서 파일을 한 번 열어 저장한 뒤 "
            f"다시 시도해 주세요. 문제 위치: {preview}{more}",
        )
    return [
        f"{context}의 계산되지 않은 수식 셀 {len(labels)}개를 빈 값으로 두고 넘어갔습니다. "
        f"해당 항목은 집계에서 빠집니다. 문제 위치: {preview}{more}"
    ]


def _pick_sheet(workbook, sheet_name: Optional[str], path: str):
    if not sheet_name:
        return workbook[workbook.sheetnames[0]]
    if sheet_name not in workbook.sheetnames:
        raise SheetNotFoundError(
            f"'{sheet_name}' 시트가 없습니다.",
            "이 파일의 시트: " + ", ".join(workbook.sheetnames),
        )
    return workbook[sheet_name]


def _merged_lookup(sheet) -> dict[tuple[int, int], tuple[int, int]]:
    """병합 셀의 모든 좌표를 '대표 셀(좌상단)' 좌표로 연결하는 표."""
    lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in sheet.merged_cells.ranges:
        anchor = (rng.min_row, rng.min_col)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                lookup[(r, c)] = anchor
    return lookup


def _value_at(sheet, merged, row: int, col: int) -> Any:
    anchor = merged.get((row, col))
    if anchor is not None:
        row, col = anchor
    return sheet.cell(row=row, column=col).value


def _read_grid(sheet, merged, row1: int, col1: int, row2: int, col2: int) -> list[list[Any]]:
    grid: list[list[Any]] = []
    for r in range(row1, row2 + 1):
        grid.append([_value_at(sheet, merged, r, c) for c in range(col1, col2 + 1)])
    return grid


def _detect_bounds(sheet, merged) -> Optional[tuple[int, int, int, int]]:
    """시트에서 '진짜 표'가 시작하는 위치를 추정한다.

    맨 위 몇 줄이 제목/작성일 같은 안내문인 경우가 흔하므로, 채워진 셀이
    2개 이상인 첫 행을 헤더 시작으로 본다.

    시작 위치와 컬럼 범위는 위쪽 ``_SCAN_ROWS`` 행만 보고 정하지만, **끝 행은
    시트 전체**를 기준으로 잡는다. 그러지 않으면 수백 행짜리 표가 잘린다.
    """
    sheet_max_row = sheet.max_row or 0
    max_col = min(sheet.max_column or 0, _SCAN_COLS)
    scan_last = min(sheet_max_row, _SCAN_ROWS)
    if sheet_max_row < 1 or max_col < 1:
        return None

    filled: list[tuple[int, list[int]]] = []
    for r in range(1, scan_last + 1):
        cols = [c for c in range(1, max_col + 1) if _is_filled(_value_at(sheet, merged, r, c))]
        if cols:
            filled.append((r, cols))
    if not filled:
        return None

    start_row = filled[0][0]
    for r, cols in filled:
        if len(cols) >= 2:
            start_row = r
            break

    body = [item for item in filled if item[0] >= start_row]
    if not body:
        return None
    min_col = min(min(cols) for _, cols in body)
    max_col_used = max(max(cols) for _, cols in body)

    # 끝 행은 시트 전체 기준. 뒤쪽의 빈 행(서식만 남은 행)은 잘라낸다.
    end_row = sheet_max_row
    while end_row > start_row and not any(
        _is_filled(_value_at(sheet, merged, end_row, c))
        for c in range(min_col, max_col_used + 1)
    ):
        end_row -= 1
    return start_row, min_col, end_row, max_col_used


def _detect_header_rows(grid: Sequence[Sequence[Any]], fallback: int) -> int:
    """헤더가 몇 행인지 추정한다.

    첫 행에 빈 칸이 있고 둘째 행이 꽉 차 있으면 '병합된 상위 헤더 + 하위 헤더'
    구조로 보고 2행을 헤더로 잡는다. (병합 셀은 이미 채워져 들어오므로
    '상위 헤더가 같은 값으로 반복되는지'로 판단한다.)
    """
    if len(grid) < 3:
        return min(fallback, len(grid))

    first = list(grid[0])
    second = list(grid[1])
    third = list(grid[2])

    first_texts = [_is_texty(v) for v in first]
    second_texts = [_is_texty(v) for v in second]
    third_texts = [_is_texty(v) for v in third]

    # 2행까지 전부 글자인데 3행부터 숫자/날짜가 섞이면 2단 헤더로 본다.
    # (병합된 상위 헤더는 이미 같은 값으로 채워져 들어오므로 글자로 잡힌다.)
    if all(first_texts) and all(second_texts) and not all(third_texts):
        return 2
    return min(fallback, len(grid))


def _build_columns(header_grid: Sequence[Sequence[Any]], first_col: int) -> list[str]:
    if not header_grid:
        return []
    width = max(len(row) for row in header_grid)
    columns: list[str] = []
    for c in range(width):
        parts: list[str] = []
        for row in header_grid:
            value = row[c] if c < len(row) else None
            text = _stringify(value)
            if text and (not parts or parts[-1] != text):
                parts.append(text)
        name = " / ".join(parts).strip()
        if not name:
            name = f"열{get_column_letter(first_col + c)}"
        columns.append(name)
    return columns


def _dedupe(names: Iterable[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for raw in names:
        name = _stringify(raw) or "이름없음"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.split())
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _is_filled(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _is_texty(value: Any) -> bool:
    """헤더 후보인지 판단: 비어 있거나 문자열이면 True."""
    if value is None:
        return True
    if isinstance(value, str):
        return True
    return False


def coordinate(text: str) -> tuple[int, int]:
    """``'B3'`` -> ``(3, 2)``."""
    try:
        col_letter, row = coordinate_from_string(text.strip().upper())
        return row, column_index_from_string(col_letter)
    except Exception as exc:
        raise CellRangeError(
            f"셀 좌표 '{text}' 를 이해할 수 없습니다.",
            "'B3' 처럼 [열문자][행번호] 형식으로 입력해 주세요.",
        ) from exc
