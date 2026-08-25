"""아주 단순한 엑셀 수식 계산기.

'계산된 값 캐시가 없는' 엑셀 파일(프로그램이 만들었거나, 재계산 없이
저장된 경우)에서 자주 나오는 간단한 수식 — 셀 참조 덧셈/뺄셈(하루씩
더해가는 날짜 헤더 등), SUM/AVERAGE/MAX/MIN/COUNT, IF 비교 — 을 대신
계산해서 값을 채운다.

엑셀의 모든 함수를 지원하지는 않는다. 그럴 필요도 없을뿐더러, 잘못
계산해서 '그럴듯하지만 틀린 값'을 내놓는 것이 "이 셀은 계산 못 했다"고
솔직하게 말하는 것보다 훨씬 위험하다. 그래서 지원 범위 밖의 수식은
조용히 포기하고(``UNRESOLVED``) 상위 코드가 기존 안전장치
(:class:`~reportgen.errors.FormulaCacheError` 또는 '빈 값으로 넘어가기'
옵션)로 처리하게 넘긴다.

보안 메모: 이 모듈은 ``eval()`` 을 쓰지 않는다. 수식 텍스트를 파이썬
문법으로 옮긴 뒤 :mod:`ast` 로 파싱하고, 허용한 노드 종류만 손수 순회해서
계산한다. 화이트리스트에 없는 노드/함수를 만나면 예외를 던지고 즉시
포기한다.
"""

from __future__ import annotations

import ast
import datetime as _dt
import re
from typing import Any, Optional

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

__all__ = ["FormulaEngine", "UNRESOLVED"]

#: 계산을 포기했다는 표시. ``None`` 과 구분해야 한다 — 진짜 빈 셀은 ``None``
#: 그대로 두고 아무 경고도 내지 않아야 하기 때문이다.
UNRESOLVED = object()

_ALLOWED_FUNCS = frozenset({"SUM", "AVERAGE", "MAX", "MIN", "COUNT", "IF", "ROUND", "ABS"})
_CELL_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d+)$")
_RANGE_TOKEN_RE = re.compile(
    r"\$?[A-Za-z]{1,3}\$?\d+\s*:\s*\$?[A-Za-z]{1,3}\$?\d+"
)
#: 파이썬 소스 문자열은 NUL 바이트를 담을 수 없으므로, 눈에 띄지만 평범한
#: 수식에는 나오지 않을 표식 문자열을 쓴다.
_RANGE_MARK = "@@RANGE@@"
_NE_MARK = "@@NE@@"
_MAX_RANGE_CELLS = 5000  # 실수로 시트 전체를 참조해도 멈춰 있지 않도록


class _GiveUp(Exception):
    """지원하지 않는 수식 형태를 만나면 여기서 끊는다."""


class FormulaEngine:
    """한 시트 안에서 '캐시가 없는 수식 셀'을 대신 계산한다.

    ``data_sheet`` 는 ``data_only=True`` 로 연 워크북의 시트(계산된 값을
    읽음), ``formula_sheet`` 는 ``data_only=False`` 로 연 같은 시트(수식
    텍스트를 읽음)다. 캐시된 값이 있으면 그대로 쓰고, 없을 때만 수식을
    직접 계산한다 — 그래서 정상적인 대부분의 파일에서는 이 계산기가
    사실상 관여하지 않는다.
    """

    def __init__(self, data_sheet, formula_sheet, merged: dict[tuple[int, int], tuple[int, int]]) -> None:
        self.data_sheet = data_sheet
        self.formula_sheet = formula_sheet
        self.merged = merged
        self._cache: dict[tuple[int, int], Any] = {}
        self._evaluating: set[tuple[int, int]] = set()
        #: 계산을 포기한 (수식이었던) 셀 좌표들. 상위 코드가 이 목록으로
        #: FormulaCacheError / 경고 메시지를 만든다.
        self.unresolved: set[tuple[int, int]] = set()

    # ------------------------------------------------------------------ #
    def get(self, row: int, col: int) -> Any:
        """``(row, col)`` 셀의 값을 얻는다. 계산 못 하면 ``None``."""
        key = (row, col)
        anchor = self.merged.get(key)
        if anchor is not None and anchor != key:
            return self.get(*anchor)

        if key in self._cache:
            cached = self._cache[key]
            return None if cached is UNRESOLVED else cached

        cached_value = self.data_sheet.cell(row=row, column=col).value
        if cached_value is not None:
            self._cache[key] = cached_value
            return cached_value

        formula_cell = self.formula_sheet.cell(row=row, column=col)
        if formula_cell.data_type != "f":
            self._cache[key] = None  # 진짜 빈 셀
            return None

        if key in self._evaluating:
            self._cache[key] = UNRESOLVED  # 순환 참조 - 포기
            self.unresolved.add(key)
            return None

        self._evaluating.add(key)
        try:
            value = self._evaluate(formula_cell.value)
        except _GiveUp:
            value = UNRESOLVED
        finally:
            self._evaluating.discard(key)

        self._cache[key] = value
        if value is UNRESOLVED:
            self.unresolved.add(key)
            return None
        return value

    def resolve_cell(self, name: str) -> Any:
        if name.upper() == "TRUE":
            return True
        if name.upper() == "FALSE":
            return False
        match = _CELL_RE.match(name)
        if not match:
            raise _GiveUp(f"셀 참조가 아님: {name}")
        col = column_index_from_string(match.group(1).upper())
        row = int(match.group(2))
        return self.get(row, col)

    def resolve_range(self, range_text: str) -> list[Any]:
        try:
            start, end = range_text.split(":")
            row1, col1 = _coord(start)
            row2, col2 = _coord(end)
        except Exception as exc:  # noqa: BLE001
            raise _GiveUp(f"범위를 이해할 수 없음: {range_text}") from exc
        row1, row2 = sorted((row1, row2))
        col1, col2 = sorted((col1, col2))
        if (row2 - row1 + 1) * (col2 - col1 + 1) > _MAX_RANGE_CELLS:
            raise _GiveUp("범위가 너무 큼")
        values = []
        for r in range(row1, row2 + 1):
            for c in range(col1, col2 + 1):
                values.append(self.get(r, c))
        return values

    # ------------------------------------------------------------------ #
    def _evaluate(self, formula_text: str) -> Any:
        expr = _translate(formula_text)
        try:
            tree = ast.parse(expr, mode="eval")
        except SyntaxError as exc:
            raise _GiveUp(f"수식 문법을 이해할 수 없음: {formula_text}") from exc
        return _eval_node(tree, self)


# --------------------------------------------------------------------------- #
# 엑셀 수식 텍스트 -> 파이썬 표현식
# --------------------------------------------------------------------------- #
def _translate(formula: str) -> str:
    text = formula[1:] if formula.startswith("=") else formula

    def wrap_range(match: "re.Match[str]") -> str:
        return f'"{_RANGE_MARK}{match.group(0).replace(" ", "")}"'

    text = _RANGE_TOKEN_RE.sub(wrap_range, text)
    text = text.replace("<>", _NE_MARK)
    # 비교의 '=' 를 '==' 로 (이미 <=, >=, == 인 경우는 건드리지 않는다)
    text = re.sub(r"(?<![<>=!\x00])=(?!=)", "==", text)
    text = text.replace(_NE_MARK, "!=")
    text = text.replace("^", "**")
    return text


def _coord(cell_ref: str) -> tuple[int, int]:
    letter, row = coordinate_from_string(cell_ref.replace("$", "").strip())
    return int(row), column_index_from_string(letter)


# --------------------------------------------------------------------------- #
# 제한된 AST 평가기 (eval() 을 쓰지 않는다)
# --------------------------------------------------------------------------- #
def _eval_node(node: ast.AST, engine: FormulaEngine) -> Any:
    if isinstance(node, ast.Expression):
        return _eval_node(node.body, engine)

    if isinstance(node, ast.Constant):
        value = node.value
        if isinstance(value, str) and value.startswith(_RANGE_MARK):
            return engine.resolve_range(value[len(_RANGE_MARK) :])
        return value

    if isinstance(node, ast.Name):
        return engine.resolve_cell(node.id)

    if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
        return -_num(_eval_node(node.operand, engine))

    if isinstance(node, ast.BinOp):
        left = _eval_node(node.left, engine)
        right = _eval_node(node.right, engine)
        return _apply_binop(node.op, left, right)

    if isinstance(node, ast.Compare) and len(node.ops) == 1 and len(node.comparators) == 1:
        left = _eval_node(node.left, engine)
        right = _eval_node(node.comparators[0], engine)
        return _apply_compare(node.ops[0], left, right)

    if isinstance(node, ast.Call):
        return _eval_call(node, engine)

    raise _GiveUp(f"지원하지 않는 수식 구조: {ast.dump(node)}")


def _eval_call(node: ast.Call, engine: FormulaEngine) -> Any:
    if not isinstance(node.func, ast.Name) or node.func.id.upper() not in _ALLOWED_FUNCS:
        raise _GiveUp("지원하지 않는 함수")
    name = node.func.id.upper()

    if name == "IF":
        if not (2 <= len(node.args) <= 3):
            raise _GiveUp("IF 인자 개수")
        condition = _eval_node(node.args[0], engine)
        if condition:
            return _eval_node(node.args[1], engine)
        return _eval_node(node.args[2], engine) if len(node.args) == 3 else False

    args = [_eval_node(a, engine) for a in node.args]
    flat: list[Any] = []
    for value in args:
        if isinstance(value, list):
            flat.extend(value)
        else:
            flat.append(value)

    if name == "ROUND":
        if len(args) != 2:
            raise _GiveUp("ROUND 인자 개수")
        return round(_num(args[0]), int(args[1]))
    if name == "ABS":
        if len(args) != 1:
            raise _GiveUp("ABS 인자 개수")
        return abs(_num(args[0]))

    numbers = [v for v in flat if isinstance(v, (int, float)) and not isinstance(v, bool)]
    if name == "SUM":
        return sum(numbers)
    if name == "AVERAGE":
        if not numbers:
            raise _GiveUp("AVERAGE 대상 없음")
        return sum(numbers) / len(numbers)
    if name == "MAX":
        return max(numbers) if numbers else 0
    if name == "MIN":
        return min(numbers) if numbers else 0
    if name == "COUNT":
        return len(numbers)
    raise _GiveUp(f"미구현 함수: {name}")  # pragma: no cover - _ALLOWED_FUNCS 와 항상 동기화됨


def _apply_binop(op: ast.operator, left: Any, right: Any) -> Any:
    if isinstance(op, ast.Add):
        if isinstance(left, _dt.date) and not isinstance(right, _dt.date):
            return left + _dt.timedelta(days=_num(right))
        if isinstance(right, _dt.date) and not isinstance(left, _dt.date):
            return right + _dt.timedelta(days=_num(left))
        return _num(left) + _num(right)
    if isinstance(op, ast.Sub):
        if isinstance(left, _dt.date) and isinstance(right, _dt.date):
            return (left - right).days
        if isinstance(left, _dt.date) and not isinstance(right, _dt.date):
            return left - _dt.timedelta(days=_num(right))
        return _num(left) - _num(right)
    if isinstance(op, ast.Mult):
        return _num(left) * _num(right)
    if isinstance(op, ast.Div):
        divisor = _num(right)
        if divisor == 0:
            raise _GiveUp("0으로 나눔")
        return _num(left) / divisor
    if isinstance(op, ast.Pow):
        return _num(left) ** _num(right)
    raise _GiveUp("지원하지 않는 연산자")


def _apply_compare(op: ast.cmpop, left: Any, right: Any) -> bool:
    if isinstance(op, ast.Gt):
        return left > right
    if isinstance(op, ast.Lt):
        return left < right
    if isinstance(op, ast.GtE):
        return left >= right
    if isinstance(op, ast.LtE):
        return left <= right
    if isinstance(op, ast.Eq):
        return left == right
    if isinstance(op, ast.NotEq):
        return left != right
    raise _GiveUp("지원하지 않는 비교 연산자")


def _num(value: Any) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise _GiveUp(f"숫자가 아님: {value!r}")
    return value
