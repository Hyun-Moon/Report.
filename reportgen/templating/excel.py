"""엑셀(.xlsx/.xlsm) 템플릿 처리 - openpyxl 기반.

두 가지 방식을 함께 지원한다.

1. **플레이스홀더 방식** - 템플릿 셀에 ``{{사용량}}`` 이라고 적어 두면 그 셀을
   찾아서 값으로 바꾼다. 셀 위치가 바뀌어도 태그만 따라가므로 관리가 편하다.
2. **셀 좌표 방식** - 매핑 키를 ``Sheet1!B3`` 처럼 직접 지정한다. 이미 만들어진
   양식을 손대지 않고 값만 꽂아 넣을 때 쓴다.

서식(테두리/색/글꼴/표시형식)과 수식은 ``.value`` 만 바꾸므로 그대로 남는다.
"""

from __future__ import annotations

import datetime as _dt
import os
import re
from typing import Any, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..data_reader import coordinate
from ..errors import TemplateError
from ..mapping import TemplateSlot
from .base import TABLE_ANCHOR_RE, TAG_RE, TemplateHandler, is_simple_tag

__all__ = ["ExcelTemplate", "CELL_KEY_RE"]

#: ``Sheet1!B3`` 형태의 직접 좌표 매핑 키
CELL_KEY_RE = re.compile(r"^\s*(?:'([^']+)'|([^!]+))!\s*([A-Za-z]{1,3}\d+)\s*$")

# 스캔 상한 (깨진 파일에서 무한정 도는 것을 막는다)
_MAX_ROW = 2000
_MAX_COL = 200


class ExcelTemplate(TemplateHandler):
    kind = "excel"
    extension = ".xlsx"

    def __init__(self, path: str) -> None:
        super().__init__(path)
        if os.path.splitext(path)[1].lower() == ".xlsm":
            self.extension = ".xlsm"
        self._slots: Optional[list[TemplateSlot]] = None

    # ------------------------------------------------------------------ #
    # 스캔
    # ------------------------------------------------------------------ #
    def scan(self) -> list[TemplateSlot]:
        if self._slots is not None:
            return self._slots

        workbook = self._load()
        try:
            found: dict[str, TemplateSlot] = {}
            for sheet in workbook.worksheets:
                for cell_ref, text in _iter_text_cells(sheet):
                    for match in TAG_RE.finditer(text):
                        inner = match.group(1).strip()
                        if TABLE_ANCHOR_RE.match(inner):
                            key = "#표"
                            slot_kind = "table"
                        elif is_simple_tag(inner):
                            key = inner
                            slot_kind = "tag"
                        else:
                            continue
                        where = f"{sheet.title}!{cell_ref}"
                        slot = found.get(key)
                        if slot is None:
                            found[key] = TemplateSlot(
                                key=key,
                                kind=slot_kind,
                                where=where,
                                sample=text.strip()[:60],
                                occurrences=1,
                            )
                        else:
                            slot.occurrences += 1
                            if where not in slot.where:
                                slot.where = f"{slot.where}, {where}"
            self._slots = list(found.values())
            return self._slots
        finally:
            workbook.close()

    def sheet_names(self) -> list[str]:
        workbook = self._load()
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    # ------------------------------------------------------------------ #
    # 렌더링
    # ------------------------------------------------------------------ #
    def render(
        self,
        context: dict[str, Any],
        output_path: str,
        table_data: Optional[dict] = None,
    ) -> str:
        workbook = self._load()
        try:
            # 1) 셀 좌표를 직접 지정한 항목 (Sheet1!B3 형태)
            direct = {k: v for k, v in context.items() if CELL_KEY_RE.match(k)}
            tags = {k: v for k, v in context.items() if k not in direct}

            for key, value in direct.items():
                self._write_direct(workbook, key, value)

            # 2) 플레이스홀더 치환
            for sheet in workbook.worksheets:
                for cell_ref, text in _iter_text_cells(sheet):
                    if TAG_RE.search(text) is None:
                        continue
                    cell = sheet[cell_ref]
                    if isinstance(cell, MergedCell):
                        continue
                    anchor = _table_anchor(text)
                    if anchor:
                        _write_table(sheet, cell_ref, table_data)
                        continue
                    cell.value = _substitute(text, tags)

            _mark_recalculate(workbook)

            os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
            try:
                workbook.save(output_path)
            except PermissionError as exc:
                raise TemplateError(
                    f"결과 파일을 저장하지 못했습니다: {os.path.basename(output_path)}",
                    "같은 이름의 파일이 엑셀에서 열려 있으면 닫은 뒤 다시 시도해 주세요.",
                ) from exc
            return os.path.abspath(output_path)
        finally:
            workbook.close()

    # ------------------------------------------------------------------ #
    # 내부
    # ------------------------------------------------------------------ #
    def _load(self):
        try:
            return load_workbook(
                self.path,
                data_only=False,  # 수식을 문자열 그대로 보존
                keep_vba=self.path.lower().endswith(".xlsm"),
            )
        except Exception as exc:  # noqa: BLE001
            raise TemplateError(
                f"엑셀 템플릿을 열지 못했습니다: {self.name}",
                f"파일이 손상되었거나 암호가 걸려 있을 수 있습니다. ({exc})",
            ) from exc

    def _write_direct(self, workbook, key: str, value: Any) -> None:
        match = CELL_KEY_RE.match(key)
        if not match:
            return
        sheet_name = (match.group(1) or match.group(2) or "").strip()
        cell_ref = match.group(3).upper()
        if sheet_name not in workbook.sheetnames:
            raise TemplateError(
                f"매핑에 적힌 '{sheet_name}' 시트가 템플릿에 없습니다.",
                "템플릿의 시트: " + ", ".join(workbook.sheetnames),
            )
        sheet = workbook[sheet_name]
        row, col = coordinate(cell_ref)
        target = sheet.cell(row=row, column=col)
        if isinstance(target, MergedCell):
            target = _merged_anchor(sheet, row, col)
        target.value = _coerce(value)


def _iter_text_cells(sheet: Worksheet) -> list[tuple[str, str]]:
    """문자열이 들어 있는 셀만 (좌표, 문자열) 로 돌려준다."""
    out: list[tuple[str, str]] = []
    max_row = min(sheet.max_row or 0, _MAX_ROW)
    max_col = min(sheet.max_column or 0, _MAX_COL)
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                out.append((cell.coordinate, cell.value))
    return out


def _table_anchor(text: str) -> bool:
    for match in TAG_RE.finditer(text):
        if TABLE_ANCHOR_RE.match(match.group(1).strip()):
            return True
    return False


def _substitute(text: str, values: dict[str, Any]) -> Any:
    """셀 문자열의 태그를 값으로 바꾼다.

    셀 전체가 태그 하나뿐이면 원래 자료형(숫자/날짜)을 유지한 채 넣는다.
    그래야 셀에 걸린 표시 형식(#,##0 등)과 이후 수식이 제대로 동작한다.
    """
    matches = list(TAG_RE.finditer(text))
    if len(matches) == 1 and matches[0].group(0).strip() == text.strip():
        key = matches[0].group(1).strip()
        if key in values:
            return _coerce(values[key])
        # 매핑되지 않은 태그는 비운다. 다만 태그처럼 안 생긴 내용(사용자가 그냥
        # 적어 둔 중괄호 문구)은 지우지 않고 그대로 남긴다.
        return "" if is_simple_tag(key) else text

    def replace(match: "re.Match[str]") -> str:
        key = match.group(1).strip()
        if key in values:
            value = values[key]
            return "" if value is None else str(value)
        return "" if is_simple_tag(key) else match.group(0)

    return TAG_RE.sub(replace, text)


def _coerce(value: Any) -> Any:
    """openpyxl 이 셀에 넣을 수 있는 자료형으로 맞춘다."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool, str, _dt.date, _dt.datetime, _dt.time)):
        return value
    return str(value)


def _merged_anchor(sheet: Worksheet, row: int, col: int):
    for rng in sheet.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return sheet.cell(row=rng.min_row, column=rng.min_col)
    return sheet.cell(row=row, column=col)


def _write_table(sheet: Worksheet, anchor_ref: str, table_data: Optional[dict]) -> None:
    """``{{#표}}`` 자리부터 아래로 표를 써 내려간다."""
    row, col = coordinate(anchor_ref)
    data = table_data or {}
    columns: list[str] = list(data.get("columns") or [])
    rows: Iterable[Iterable[Any]] = data.get("rows") or []
    include_header = bool(data.get("include_header", True))

    sheet.cell(row=row, column=col).value = None

    cursor = row
    if include_header and columns:
        for offset, name in enumerate(columns):
            _safe_write(sheet, cursor, col + offset, name)
        cursor += 1
    for record in rows:
        for offset, value in enumerate(record):
            _safe_write(sheet, cursor, col + offset, _coerce(value))
        cursor += 1


def _safe_write(sheet: Worksheet, row: int, col: int, value: Any) -> None:
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        cell = _merged_anchor(sheet, row, col)
    cell.value = value


def _mark_recalculate(workbook) -> None:
    """열 때 수식을 다시 계산하도록 표시한다.

    openpyxl 은 수식의 '계산된 값' 캐시를 보존하지 못하므로, 이 표시가 없으면
    수식 셀이 0 또는 빈 값으로 보일 수 있다.
    """
    try:
        workbook.calculation.fullCalcOnLoad = True
    except AttributeError:  # pragma: no cover - openpyxl 구버전 대비
        pass


def format_cell_key(sheet_name: str, cell_ref: str) -> str:
    """``('실적', 'B3')`` -> ``"'실적'!B3"`` (공백/느낌표 대비)."""
    if re.search(r"[\s!']", sheet_name):
        return f"'{sheet_name}'!{cell_ref.upper()}"
    return f"{sheet_name}!{cell_ref.upper()}"


def cell_label(sheet_name: str, row: int, col: int) -> str:
    return format_cell_key(sheet_name, f"{get_column_letter(col)}{row}")
