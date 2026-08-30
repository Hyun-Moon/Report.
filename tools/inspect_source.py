#!/usr/bin/env python3
"""원본 엑셀의 '실제 구조'를 있는 그대로 출력한다.

원본 파일을 남에게 보내지 않고도 "이 프로그램이 그 파일을 어떻게 보고 있는지"
확인·공유하기 위한 도구다. 추측이나 요약 없이 openpyxl 이 읽은 것만 적는다.

기본 사용::

    python tools/inspect_source.py "원본.xlsx"

값이 민감해서 밖으로 내보내면 안 될 때 (숫자는 자릿수만, 글자는 길이만 표시)::

    python tools/inspect_source.py "원본.xlsx" --mask

결과를 파일로 저장해서 붙여넣기 편하게::

    python tools/inspect_source.py "원본.xlsx" --mask > 구조.txt
"""

from __future__ import annotations

import argparse
import datetime as _dt
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from reportgen.data_reader import list_table_blocks  # noqa: E402

#: 시트마다 표 후보를 최대 몇 개까지 보여줄지
_MAX_BLOCKS = 12
#: 표 후보마다 미리보기 행을 몇 줄까지 보여줄지
_MAX_PREVIEW_ROWS = 6
#: 한 줄에 보여줄 칸 수
_MAX_PREVIEW_COLS = 10


def mask_value(value, masked: bool) -> str:
    """셀 값을 사람이 읽을 수 있게 바꾼다. ``masked`` 면 실제 값을 숨긴다."""
    if value is None:
        return "·"
    if isinstance(value, str):
        text = " ".join(value.split())
        if not text:
            return "·"
        if masked:
            return f"<글자{len(text)}자>"
        return text if len(text) <= 22 else text[:21] + "…"
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        if masked:
            return f"<숫자{len(str(abs(int(value))))}자리>"
        return str(value)
    if isinstance(value, (_dt.datetime, _dt.date)):
        # 날짜는 가려도 안전한 편이고, 날짜 인식이 되는지가 중요하므로 형태를 남긴다.
        return value.strftime("%Y-%m-%d")
    return f"<{type(value).__name__}>"


def describe_workbook(path: str, masked: bool) -> None:
    print("=" * 78)
    print(f"파일: {os.path.basename(path)}")
    print(f"크기: {os.path.getsize(path):,} 바이트")
    print(f"값 가리기: {'켬 (실제 값 안 보임)' if masked else '끔'}")
    print("=" * 78)

    # data_only=True  -> 수식의 '계산된 값'
    # data_only=False -> 수식 문자열 그 자체
    values_wb = load_workbook(path, data_only=True)
    formula_wb = load_workbook(path, data_only=False)

    try:
        print(f"\n시트 {len(values_wb.sheetnames)}개: {', '.join(values_wb.sheetnames)}")

        for name in values_wb.sheetnames:
            describe_sheet(path, values_wb[name], formula_wb[name], masked)
    finally:
        values_wb.close()
        formula_wb.close()


def describe_sheet(path: str, sheet, formula_sheet, masked: bool) -> None:
    print()
    print("-" * 78)
    print(f"[시트] {sheet.title}")
    print("-" * 78)
    print(f"  사용 범위: {sheet.dimensions}  (최대 {sheet.max_row}행 x {sheet.max_column}열)")

    merged = list(sheet.merged_cells.ranges)
    if merged:
        shown = ", ".join(str(r) for r in merged[:10])
        more = f" 외 {len(merged) - 10}개" if len(merged) > 10 else ""
        print(f"  병합 셀 {len(merged)}개: {shown}{more}")
    else:
        print("  병합 셀 없음")

    # 계산된 값이 없는 수식 셀 (이게 있으면 읽을 때 문제가 된다)
    stale = []
    for row in formula_sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                cached = sheet[cell.coordinate].value
                if cached is None:
                    stale.append(cell.coordinate)
                if len(stale) >= 200:
                    break
    if stale:
        print(
            f"  ⚠ 계산된 값이 저장 안 된 수식 셀 {len(stale)}개"
            f"{'+' if len(stale) >= 200 else ''}: {', '.join(stale[:8])}"
        )

    # 수식 몇 개를 실제 문자열 그대로 (어떤 함수를 쓰는지 확인용)
    formulas = []
    for row in formula_sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f" and isinstance(cell.value, str):
                formulas.append((cell.coordinate, cell.value))
            if len(formulas) >= 8:
                break
        if len(formulas) >= 8:
            break
    if formulas:
        print("  수식 예시 (실제 저장된 문자열):")
        for coord, text in formulas:
            body = text if len(text) <= 60 else text[:59] + "…"
            print(f"    {coord} : {body}")

    # 이 프로그램이 이 시트에서 '표'로 인식하는 덩어리들
    try:
        blocks = list_table_blocks(path, sheet.title)
    except Exception as exc:  # noqa: BLE001 - 진단 도구이므로 계속 진행
        print(f"  표 후보 찾기 실패: {type(exc).__name__}: {exc}")
        return

    if not blocks:
        print("  표 후보: 없음 (빈 시트이거나 내용이 흩어져 있음)")
        return

    print(f"  표 후보 {len(blocks)}개:")
    for block in blocks[:_MAX_BLOCKS]:
        print(f"    · {block.range}  ({block.n_rows}행 x {block.n_cols}열)  {block.preview}")
    if len(blocks) > _MAX_BLOCKS:
        print(f"    … 외 {len(blocks) - _MAX_BLOCKS}개")

    # 가장 큰 표 후보의 내용을 몇 줄 보여준다 (행 번호 포함 - 취합 기능에서 필요)
    biggest = max(blocks, key=lambda b: b.n_rows * b.n_cols)
    print(f"\n  가장 큰 표 후보 {biggest.range} 미리보기:")
    _print_block(sheet, biggest, masked)


def _print_block(sheet, block, masked: bool) -> None:
    last_col = min(block.col2, block.col1 + _MAX_PREVIEW_COLS - 1)
    header = "        " + " | ".join(
        f"{get_column_letter(c):>10}" for c in range(block.col1, last_col + 1)
    )
    print(header)
    last_row = min(block.row2, block.row1 + _MAX_PREVIEW_ROWS - 1)
    for r in range(block.row1, last_row + 1):
        cells = []
        for c in range(block.col1, last_col + 1):
            cells.append(f"{mask_value(sheet.cell(row=r, column=c).value, masked):>10}")
        print(f"  {r:>4} | " + " | ".join(cells))
    if block.row2 > last_row:
        print(f"       … 아래로 {block.row2 - last_row}행 더 있음")
    if block.col2 > last_col:
        print(f"       … 오른쪽으로 {block.col2 - last_col}열 더 있음")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="원본 엑셀의 실제 구조를 출력한다.")
    parser.add_argument("path", help="살펴볼 .xlsx / .xlsm 파일")
    parser.add_argument(
        "--mask",
        action="store_true",
        help="실제 값을 숨기고 자릿수/글자수만 표시 (밖으로 공유할 때)",
    )
    args = parser.parse_args(argv)

    if not os.path.isfile(args.path):
        print(f"파일을 찾을 수 없습니다: {args.path}", file=sys.stderr)
        return 1
    try:
        describe_workbook(args.path, args.mask)
    except Exception as exc:  # noqa: BLE001 - 진단 도구이므로 원인을 그대로 보여준다
        print(f"\n읽는 중 오류: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
